# scripts/generate_qa_checklists.py
# Generates ONE Excel file with 3 tabs: Summary, Entry Criteria, Exit Criteria (QA-only)

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


# -----------------------------
# Output path (always root/qa/)
# -----------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
QA_DIR = PROJECT_ROOT / "qa"
timestamp_for_file = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
OUTPUT_PATH = QA_DIR / f"qa_checklists_{timestamp_for_file}.xlsx"

generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# -----------------------------
# Status dropdown + colours
# -----------------------------
STATUS_OPTIONS = ["Open", "Completed", "Pending", "Blocked"]

FILL_OPEN = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")        # Blue
FILL_DONE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")        # Green
FILL_PENDING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # Yellow
FILL_BLOCKED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Red


# -----------------------------
# Entry Criteria (Dev + QA)
# -----------------------------
ENTRY_DEV_CHECKLIST = [
    "Requirements implemented as per agreed scope",
    "Acceptance criteria addressed",
    "Edge cases and non-functional expectations considered",
    "Unit testing completed",
    "Unit test results reviewed and recorded",
    "No failing unit tests",
    "Code review completed",
    "Change log updated",
    "Build deployed successfully to test environment",
    "Environment configuration completed",
    "Access and permissions provided",
    "No critical build or deployment issues",
    "Dependencies from other teams ready",
    "No merge conflicts affecting the build",
    "Security considerations reviewed",
    "Data privacy rules reviewed",
    "Logging and monitoring enabled",
    "Feature flags configured if applicable",
    "Developer confirms build is ready for QA testing",
]

ENTRY_QA_CHECKLIST = [
    "Entry criteria document reviewed",
    "Requirements and acceptance criteria understood",
    "Test scope confirmed",
    "Test plan or checklist prepared",
    "Test approach reviewed",
    "Risks and assumptions identified",
    "Traceability between requirements and tests established",
    "Test data prepared and validated",
    "Boundary and negative test data prepared",
    "Test accounts available",
    "External dependencies configured or mocked",
    "Environment access verified",
    "No blocking defects present",
    "Known limitations communicated",
    "Communication channels confirmed",
    "Defect reporting guidelines understood",
    "Testing timeline agreed",
    "Prior blockers reviewed",
    "QA confirms testing can begin",
]


# -----------------------------
# Exit Criteria (QA-only) – based on your attached checklist
# + Android/iOS sections added
# -----------------------------
EXIT_SECTIONS = [
    ("1. Test Execution", [
        "All planned test cases or checklists have been executed",
        "High-risk areas have been covered",
        "Regression tests for impacted areas are completed",
    ]),
    ("2. Defect Resolution", [
        "All critical defects are fixed and retested",
        "All major defects are fixed and retested",
        "Medium and minor defects are resolved or accepted for future release",
        "No open issues block core functionality",
    ]),
    ("3. Acceptance Criteria & Quality Benchmarks", [
        "All acceptance criteria defined in requirements are met",
        "No unexpected behaviour in core workflows",
        "Performance is acceptable for this release",
        "Usability is acceptable for the team",
    ]),
    ("4. Documentation Completion", [
        "Test results are recorded and stored",
        "Known issues are documented and communicated",
        "Release notes or defect summaries are prepared (if required)",
    ]),
    ("5. Build Stability Verification", [
        "Final build is stable in the test environment",
        "No recurring crashes or environment issues",
        "Integrations with other systems/modules function correctly",
    ]),
    ("6. Stakeholder Approval", [
        "QA confirms testing is complete",
        "Product owner / team lead approves release readiness",
        "Remaining risks are communicated and accepted",
    ]),
    ("7. Mobile Release Checks – Android", [
        "Android build tested on at least one real device",
        "Android build tested on different screen sizes (where applicable)",
        "Android install/upgrade flow verified",
        "Android permissions and notifications checked (if applicable)",
    ]),
    ("8. Mobile Release Checks – iOS", [
        "iOS build tested on at least one real device",
        "iOS build tested on different screen sizes (where applicable)",
        "iOS install/upgrade flow verified",
        "iOS permissions and notifications checked (if applicable)",
    ]),
]


def add_status_conditional_formatting(ws, start_row: int, end_row: int, status_col: str):
    """Conditional formatting: colours change based on dropdown value."""
    cell_range = f"{status_col}{start_row}:{status_col}{end_row}"
    ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'{status_col}{start_row}="Open"'], fill=FILL_OPEN))
    ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'{status_col}{start_row}="Completed"'], fill=FILL_DONE))
    ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'{status_col}{start_row}="Pending"'], fill=FILL_PENDING))
    ws.conditional_formatting.add(cell_range, FormulaRule(formula=[f'{status_col}{start_row}="Blocked"'], fill=FILL_BLOCKED))


def setup_sheet_columns(ws, headers, widths):
    ws.append(headers)
    header_row = ws.max_row
    for idx, _ in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx).font = Font(bold=True)
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def generate_summary_tab(wb: Workbook):
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "QA Checklists Summary"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Generated on: {generated_at}"
    ws["A2"].font = Font(italic=True)

    # Simple fields
    fields = [
        ("Company Name", ""),
        ("Checklist Name", "Entry & Exit Criteria"),
        ("Release / Build Version", ""),
        ("Generated By (Name)", ""),
        ("Reviewed By (Dev)", ""),
        ("Reviewed By (QA)", ""),
        ("Updated/Changed By (Dev)", ""),
        ("Updated/Changed By (QA)", ""),
        ("Notes", ""),
    ]

    start_row = 4
    for i, (label, value) in enumerate(fields):
        r = start_row + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = value

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55


def generate_entry_tab(wb: Workbook, status_validation: DataValidation):
    ws = wb.create_sheet("Entry Criteria")
    ws["A1"] = "Entry Criteria Checklist"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated on: {generated_at}"
    ws["A2"].font = Font(italic=True)

    row = 4

    # Developer section
    ws[f"A{row}"] = "Developer Checklist – Build Readiness"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    setup_sheet_columns(
        ws,
        headers=["No.", "Checklist Item", "Status", "Date", "Time", "Developer"],
        widths={"A": 6, "B": 65, "C": 15, "D": 15, "E": 15, "F": 20}
    )

    dev_start = ws.max_row + 1
    for i, item in enumerate(ENTRY_DEV_CHECKLIST, start=1):
        ws.append([i, item, "Open", "", "", ""])
        status_cell = ws[f"C{ws.max_row}"]
        status_validation.add(status_cell)

    dev_end = ws.max_row

    # spacing
    ws.append([""])
    ws.append([""])
    row = ws.max_row + 1

    # QA section
    ws[f"A{row}"] = "QA Checklist – Testing Readiness"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    setup_sheet_columns(
        ws,
        headers=["No.", "Checklist Item", "Status", "Date", "Time", "QA"],
        widths={"A": 6, "B": 65, "C": 15, "D": 15, "E": 15, "F": 20}
    )

    qa_start = ws.max_row + 1
    for i, item in enumerate(ENTRY_QA_CHECKLIST, start=1):
        ws.append([i, item, "Open", "", "", ""])
        status_cell = ws[f"C{ws.max_row}"]
        status_validation.add(status_cell)

    qa_end = ws.max_row

    # Conditional formatting
    add_status_conditional_formatting(ws, dev_start, dev_end, status_col="C")
    add_status_conditional_formatting(ws, qa_start, qa_end, status_col="C")


def generate_exit_tab(wb: Workbook, status_validation: DataValidation):
    ws = wb.create_sheet("Exit Criteria")
    ws["A1"] = "Exit Criteria Checklist"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated on: {generated_at}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = "QA completion requirements before release"
    ws["A3"].font = Font(italic=True, size=10)

    # Headers (QA-only)
    ws.append(["No.", "Checklist Item", "Status", "Date", "Time", "QA"])
    header_row = ws.max_row
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws[f"{col_letter}{header_row}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20

    start_items_row = ws.max_row + 1

    item_no = 1
    for section_title, items in EXIT_SECTIONS:
        # Section row (bold, no dropdown)
        ws.append(["", section_title, "", "", "", ""])
        ws[f"B{ws.max_row}"].font = Font(bold=True)

        # Items
        for item in items:
            ws.append([item_no, item, "Open", "", "", ""])
            status_cell = ws[f"C{ws.max_row}"]
            status_validation.add(status_cell)
            item_no += 1

        ws.append([""])  # small space

    # Find the last actual item row for formatting range
    end_items_row = ws.max_row
    while end_items_row > start_items_row and ws[f"A{end_items_row}"].value in (None, ""):
        end_items_row -= 1

    # Conditional formatting for whole status column range (covers all items; section rows won’t match any status)
    add_status_conditional_formatting(ws, start_items_row, end_items_row, status_col="C")


def main():
    wb = Workbook()

    # Create dropdown validation once and reuse
    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=False
    )

    # Summary tab first
    generate_summary_tab(wb)

    # Add validation object to sheets that need dropdowns
    # (openpyxl requires adding to each worksheet)
    entry_ws = wb["Summary"]  # placeholder
    # Create Entry and Exit sheets
    generate_entry_tab(wb, status_validation)
    generate_exit_tab(wb, status_validation)

    # Add the validation object to each sheet that uses it
    wb["Entry Criteria"].add_data_validation(status_validation)
    wb["Exit Criteria"].add_data_validation(status_validation)

    QA_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Excel generated at: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
